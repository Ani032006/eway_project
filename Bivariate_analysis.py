import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent

# Load data
X_train = pd.read_csv(DATA_DIR / "X_train.csv").drop(columns=["ID"], errors="ignore")
X_test  = pd.read_csv(DATA_DIR / "X_test.csv").drop(columns=["ID"], errors="ignore")
y_train = pd.read_csv(DATA_DIR / "y_train.csv")
y_test  = pd.read_csv(DATA_DIR / "y_test.csv")

X = pd.concat([X_train, X_test], ignore_index=True)
y = pd.concat([y_train, y_test], ignore_index=True).iloc[:, 0]

binary_cols     = ['Column10','Column11','Column12','Column13','Column19','Column20','Column21']
continuous_cols = ['Column0','Column1','Column2','Column3','Column4','Column5',
                   'Column6','Column7','Column8','Column14','Column15','Column16','Column17','Column18']

total_good = (y == 0).sum()
total_bad  = (y == 1).sum()

def compute_woe_iv(grp_target):
    good = (grp_target == 0).sum()
    bad  = (grp_target == 1).sum()
    
    g_pct = good / total_good
    b_pct = bad / total_bad
    g_minus_b = g_pct - b_pct
    
    # Avoid division by zero and log of zero
    woe = np.log(max(g_pct, 1e-10) / max(b_pct, 1e-10))
    iv = woe * g_minus_b
    
    return round(g_pct * 100, 2), round(b_pct * 100, 2), round(g_minus_b * 100, 2), round(woe, 6), round(iv, 6)

# Setup Workbook
wb = Workbook()
ws = wb.active
ws.title = "Bivariate Analysis"

overall_event_rate = round((y == 0).mean() * 100, 2)
ws.append(["Overall Event Rate (%)", overall_event_rate])
ws.append([])

HEADERS = ["Variable", "Variable Category", "Frequency", "Response",
           "Response Rate (%)", "G%", "B%", "G% - B%", "WOE", "IV", "Total IV"]

# Set column widths
col_widths = [14, 22, 12, 12, 15, 10, 10, 12, 10, 10, 10]
for i, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Process each column
for col in continuous_cols + binary_cols:
    if col not in X.columns:
        continue

    combined = pd.DataFrame({"feature": X[col], "target": y})
    bucket_rows = []

    # Binning for continuous vs binary
    if col in continuous_cols:
        non_missing = combined.dropna(subset=["feature"]).copy()
        non_missing["bucket"] = pd.qcut(non_missing["feature"], q=10, duplicates="drop")
        groups = non_missing.groupby("bucket", observed=True)
    else:
        non_missing = combined.dropna(subset=["feature"]).copy()
        non_missing["bucket"] = non_missing["feature"].astype(int)
        groups = non_missing.groupby("bucket")

    # Process valid values
    for bucket, grp in groups:
        if len(grp) == 0:
            continue
            
        freq = len(grp)
        resp = int((grp["target"] == 0).sum())
        rate = round((grp["target"] == 0).mean() * 100, 2)
        g_pct, b_pct, g_minus_b, woe, iv = compute_woe_iv(grp["target"])
        
        bucket_rows.append([str(bucket), freq, resp, rate, g_pct, b_pct, g_minus_b, woe, iv])

    # Process missing values
    missing = combined[combined["feature"].isna()]
    if len(missing) > 0:
        freq = len(missing)
        resp = int((missing["target"] == 0).sum())
        rate = round((missing["target"] == 0).mean() * 100, 2)
        g_pct, b_pct, g_minus_b, woe, iv = compute_woe_iv(missing["target"])
        
        bucket_rows.append(["Missing", freq, resp, rate, g_pct, b_pct, g_minus_b, woe, iv])

    if not bucket_rows:
        continue

    # Calculate Total IV
    total_iv = round(sum(row[8] for row in bucket_rows), 6)

    # Write to Excel
    ws.append(HEADERS)
    for i, row in enumerate(bucket_rows):
        full_row = [col] + row + [total_iv if i == 0 else ""]
        ws.append(full_row)
        
    # Add some blank rows for spacing
    ws.append([])
    ws.append([])
    ws.append([])

wb.save(DATA_DIR / "bivariate_report.xlsx")
print("Saved: bivariate_report.xlsx")